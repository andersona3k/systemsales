import io
from uuid import UUID
from datetime import datetime, date
from sqlalchemy import func
from fastapi import APIRouter, Depends, HTTPException, Body, UploadFile, File, Form
from fastapi.responses import Response
from sqlalchemy.orm import Session
from ..database import get_db
from ..models.proposta import (
    PropConfig, PropTemplate, PropProduto, PropContentBlock,
    PropProposta, PropGrupo, PropItem, PropPerda, Funil, Oportunidade, OpInteracao,
)
from ..services.auth_service import get_current_user
from ..config import get_settings
from ..services.img_service import salvar_imagem

router = APIRouter(prefix="/api/proposta", tags=["Proposta"])
settings = get_settings()


# ---------- serialização ----------
def _template_out(t):
    return {"id": str(t.id), "tipo": t.tipo, "nome": t.nome, "nome_exibicao": t.nome_exibicao,
            "versao": t.versao, "ativo": t.ativo, "config": t.config or {},
            "atualizado_em": t.atualizado_em.isoformat() if t.atualizado_em else None}


def _produto_out(p):
    return {"id": str(p.id), "codigo": p.codigo, "descricao": p.descricao, "ncm": p.ncm,
            "grupo": p.grupo, "campos_extras": p.campos_extras or {}, "ativo": p.ativo}


def _bloco_out(b):
    return {"id": str(b.id), "chave": b.chave, "titulo": b.titulo, "aplica_se_a": b.aplica_se_a,
            "conteudo_padrao": b.conteudo_padrao or {}, "padrao": bool(b.padrao),
            "sugerido": bool(b.sugerido), "versao": b.versao}


def _item_out(it):
    return {"id": str(it.id), "produto_id": str(it.produto_id) if it.produto_id else None,
            "codigo_produto": it.codigo_produto, "info_adicional": it.info_adicional or "",
            "quantidade": it.quantidade or 0, "valor_unitario": it.valor_unitario or 0,
            "margem_percentual": it.margem_percentual or 0, "ordem": it.ordem or 0}


def _perda_out(pe):
    return {"id": str(pe.id), "descricao_item": pe.descricao_item, "valor_reposicao": pe.valor_reposicao or 0}


def _grupo_out(g):
    return {"id": str(g.id), "categoria": g.categoria, "moeda": g.moeda or "BRL",
            "descritivo_unidade": g.descritivo_unidade, "periodo_locacao_meses": g.periodo_locacao_meses,
            "ordem": g.ordem or 0, "itens": [_item_out(x) for x in g.itens],
            "perdas": [_perda_out(x) for x in g.perdas]}


def _proposta_out(p):
    return {"id": str(p.id), "tipo": p.tipo, "status": p.status,
            "template_id": str(p.template_id) if p.template_id else None,
            "empresa_id": str(p.empresa_id) if p.empresa_id else None,
            "pessoa_id": str(p.pessoa_id) if p.pessoa_id else None,
            "card_id": str(p.card_id) if p.card_id else None,
            "data": p.data, "validade_dias": p.validade_dias, "texto_comercial": p.texto_comercial or {},
            "moeda_padrao": p.moeda_padrao or "BRL", "blocos": p.blocos or {},
            "criado_em": p.criado_em.isoformat() if p.criado_em else None,
            "grupos": [_grupo_out(g) for g in p.grupos]}


# ---------- Config do módulo (grupos, colunas extras, etc.) ----------
@router.get("/config/{chave}")
def get_config(chave: str, db: Session = Depends(get_db), _=Depends(get_current_user)):
    c = db.query(PropConfig).filter(PropConfig.chave == chave).first()
    return {"chave": chave, "valor": c.valor if c else None}


@router.put("/config/{chave}")
def put_config(chave: str, dados: dict = Body(...), db: Session = Depends(get_db), _=Depends(get_current_user)):
    c = db.query(PropConfig).filter(PropConfig.chave == chave).first()
    if not c:
        c = PropConfig(chave=chave, valor=dados.get("valor")); db.add(c)
    else:
        c.valor = dados.get("valor")
    db.commit()
    return {"chave": chave, "valor": c.valor}


def _cols_extra(db):
    c = db.query(PropConfig).filter(PropConfig.chave == "prop_colunas").first()
    v = c.valor if c else None
    return v if isinstance(v, list) else []


# ---------- Templates (modelos) ----------
@router.get("/templates")
def listar_templates(db: Session = Depends(get_db), _=Depends(get_current_user)):
    return [_template_out(t) for t in db.query(PropTemplate).order_by(PropTemplate.tipo, PropTemplate.nome).all()]


@router.post("/templates", status_code=201)
def criar_template(dados: dict = Body(...), db: Session = Depends(get_db), _=Depends(get_current_user)):
    t = PropTemplate(tipo=dados.get("tipo") or "venda", nome=dados.get("nome") or "Novo modelo",
                     nome_exibicao=dados.get("nome_exibicao"), versao=int(dados.get("versao") or 1),
                     ativo=dados.get("ativo", True), config=dados.get("config") or {})
    db.add(t); db.commit(); db.refresh(t)
    return _template_out(t)


@router.post("/templates/{tid}/duplicar", status_code=201)
def duplicar_template(tid: UUID, dados: dict = Body(default={}), db: Session = Depends(get_db), _=Depends(get_current_user)):
    t = db.query(PropTemplate).filter(PropTemplate.id == tid).first()
    if not t: raise HTTPException(404, "Modelo nao encontrado")
    novo = PropTemplate(tipo=t.tipo, nome=(dados or {}).get("nome") or (t.nome + " (cópia)"),
                        nome_exibicao=t.nome_exibicao, versao=1, ativo=True,
                        config=dict(t.config or {}))
    db.add(novo); db.commit(); db.refresh(novo)
    return _template_out(novo)


@router.patch("/templates/{tid}")
def atualizar_template(tid: UUID, dados: dict = Body(...), db: Session = Depends(get_db), _=Depends(get_current_user)):
    t = db.query(PropTemplate).filter(PropTemplate.id == tid).first()
    if not t: raise HTTPException(404, "Modelo nao encontrado")
    for k in ("tipo", "nome", "nome_exibicao", "versao", "ativo", "config"):
        if k in dados: setattr(t, k, dados[k])
    db.commit(); db.refresh(t)
    return _template_out(t)


@router.delete("/templates/{tid}", status_code=204)
def deletar_template(tid: UUID, db: Session = Depends(get_db), _=Depends(get_current_user)):
    t = db.query(PropTemplate).filter(PropTemplate.id == tid).first()
    if t: db.delete(t); db.commit()


# ---------- Produtos (Código/Descrição/NCM/Grupo + extras) ----------
@router.get("/produtos")
def listar_produtos(db: Session = Depends(get_db), _=Depends(get_current_user)):
    return [_produto_out(p) for p in db.query(PropProduto).filter(PropProduto.ativo == True).order_by(PropProduto.grupo, PropProduto.descricao).all()]


@router.post("/produtos", status_code=201)
def criar_produto(dados: dict = Body(...), db: Session = Depends(get_db), _=Depends(get_current_user)):
    desc = (dados.get("descricao") or "").strip()
    if not desc: raise HTTPException(400, "Descricao obrigatoria")
    p = PropProduto(codigo=dados.get("codigo"), descricao=desc, ncm=dados.get("ncm"),
                    grupo=dados.get("grupo"), campos_extras=dados.get("campos_extras") or {})
    db.add(p); db.commit(); db.refresh(p)
    return _produto_out(p)


@router.patch("/produtos/{pid}")
def atualizar_produto(pid: UUID, dados: dict = Body(...), db: Session = Depends(get_db), _=Depends(get_current_user)):
    p = db.query(PropProduto).filter(PropProduto.id == pid).first()
    if not p: raise HTTPException(404, "Produto nao encontrado")
    for k in ("codigo", "descricao", "ncm", "grupo", "campos_extras", "ativo"):
        if k in dados: setattr(p, k, dados[k])
    db.commit(); db.refresh(p)
    return _produto_out(p)


@router.delete("/produtos/{pid}", status_code=204)
def deletar_produto(pid: UUID, db: Session = Depends(get_db), _=Depends(get_current_user)):
    p = db.query(PropProduto).filter(PropProduto.id == pid).first()
    if p: p.ativo = False; db.commit()


@router.get("/produtos/exportar")
def exportar_produtos(db: Session = Depends(get_db), _=Depends(get_current_user)):
    from openpyxl import Workbook
    cols = _cols_extra(db)
    wb = Workbook(); ws = wb.active; ws.title = "Produtos"
    ws.append(["Código", "Descrição", "NCM", "Grupo"] + [c.get("label", c.get("chave", "")) for c in cols])
    for p in db.query(PropProduto).filter(PropProduto.ativo == True).order_by(PropProduto.grupo, PropProduto.descricao).all():
        ex = p.campos_extras or {}
        ws.append([p.codigo, p.descricao, p.ncm, p.grupo] + [ex.get(c.get("chave", ""), "") for c in cols])
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return Response(content=buf.read(),
                    media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    headers={"Content-Disposition": "attachment; filename=produtos.xlsx"})


@router.post("/produtos/importar")
async def importar_produtos(arquivo: UploadFile = File(...), db: Session = Depends(get_db), _=Depends(get_current_user)):
    from openpyxl import load_workbook
    data = await arquivo.read()
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {"importados": 0}
    hdr = [(str(h).strip().lower() if h is not None else "") for h in rows[0]]

    def idx(*names):
        for n in names:
            if n in hdr:
                return hdr.index(n)
        return -1

    def cell(r, i):
        if i < 0 or i >= len(r) or r[i] is None:
            return None
        return str(r[i]).strip()

    ic, idd, incm, ig = idx("código", "codigo"), idx("descrição", "descricao"), idx("ncm"), idx("grupo")
    cols = _cols_extra(db)
    count = 0
    for r in rows[1:]:
        if not any(v is not None and str(v).strip() != "" for v in r):
            continue
        cod, desc = cell(r, ic), cell(r, idd)
        if not cod and not desc:
            continue
        extras = {}
        for c in cols:
            ci = idx(str(c.get("label", "")).lower())
            val = cell(r, ci)
            if val is not None:
                extras[c.get("chave", "")] = val
        p = None
        if cod:
            p = db.query(PropProduto).filter(PropProduto.codigo == cod, PropProduto.ativo == True).first()
        if p:
            if desc: p.descricao = desc
            if cell(r, incm) is not None: p.ncm = cell(r, incm)
            if cell(r, ig) is not None: p.grupo = cell(r, ig)
            if extras: p.campos_extras = {**(p.campos_extras or {}), **extras}
        else:
            db.add(PropProduto(codigo=cod, descricao=desc, ncm=cell(r, incm), grupo=cell(r, ig), campos_extras=extras))
        count += 1
    db.commit()
    return {"importados": count}


# ---------- Content Blocks (estrutura definida em Configurações) ----------
@router.get("/blocos")
def listar_blocos(db: Session = Depends(get_db), _=Depends(get_current_user)):
    return [_bloco_out(b) for b in db.query(PropContentBlock).order_by(PropContentBlock.chave).all()]


@router.post("/blocos", status_code=201)
def criar_bloco(dados: dict = Body(...), db: Session = Depends(get_db), _=Depends(get_current_user)):
    b = PropContentBlock(chave=dados.get("chave"), titulo=dados.get("titulo"),
                         aplica_se_a=dados.get("aplica_se_a") or "ambos",
                         conteudo_padrao=dados.get("conteudo_padrao") or {},
                         padrao=bool(dados.get("padrao")), sugerido=bool(dados.get("sugerido")),
                         versao=int(dados.get("versao") or 1))
    db.add(b); db.commit(); db.refresh(b)
    return _bloco_out(b)


@router.patch("/blocos/{bid}")
def atualizar_bloco(bid: UUID, dados: dict = Body(...), db: Session = Depends(get_db), _=Depends(get_current_user)):
    b = db.query(PropContentBlock).filter(PropContentBlock.id == bid).first()
    if not b: raise HTTPException(404, "Bloco nao encontrado")
    for k in ("chave", "titulo", "aplica_se_a", "conteudo_padrao", "padrao", "sugerido"):
        if k in dados: setattr(b, k, dados[k])
    if "conteudo_padrao" in dados: b.versao = (b.versao or 1) + 1
    db.commit(); db.refresh(b)
    return _bloco_out(b)


@router.delete("/blocos/{bid}", status_code=204)
def deletar_bloco(bid: UUID, db: Session = Depends(get_db), _=Depends(get_current_user)):
    b = db.query(PropContentBlock).filter(PropContentBlock.id == bid).first()
    if b: db.delete(b); db.commit()


# ---------- Propostas ----------
def _apply_grupos(db, prop_id, grupos):
    db.query(PropGrupo).filter(PropGrupo.proposta_id == prop_id).delete()
    db.flush()
    for gi, g in enumerate(grupos or []):
        grp = PropGrupo(proposta_id=prop_id, categoria=g.get("categoria"), moeda=g.get("moeda") or "BRL",
                        descritivo_unidade=g.get("descritivo_unidade"),
                        periodo_locacao_meses=g.get("periodo_locacao_meses"),
                        ordem=g.get("ordem") if g.get("ordem") is not None else gi)
        db.add(grp); db.flush()
        for ii, it in enumerate(g.get("itens") or []):
            db.add(PropItem(grupo_id=grp.id, produto_id=it.get("produto_id"), codigo_produto=it.get("codigo_produto"),
                            info_adicional=it.get("info_adicional"), quantidade=float(it.get("quantidade") or 0),
                            valor_unitario=float(it.get("valor_unitario") or 0),
                            margem_percentual=float(it.get("margem_percentual") or 0),
                            ordem=it.get("ordem") if it.get("ordem") is not None else ii))
        for pe in (g.get("perdas") or []):
            db.add(PropPerda(grupo_id=grp.id, descricao_item=pe.get("descricao_item"),
                             valor_reposicao=float(pe.get("valor_reposicao") or 0)))


@router.get("/propostas")
def listar_propostas(status: str | None = None, db: Session = Depends(get_db), _=Depends(get_current_user)):
    q = db.query(PropProposta)
    if status: q = q.filter(PropProposta.status == status)
    return [_proposta_out(p) for p in q.order_by(PropProposta.criado_em.desc()).all()]


@router.get("/propostas/{pid}")
def obter_proposta(pid: UUID, db: Session = Depends(get_db), _=Depends(get_current_user)):
    p = db.query(PropProposta).filter(PropProposta.id == pid).first()
    if not p: raise HTTPException(404, "Proposta nao encontrada")
    return _proposta_out(p)


@router.post("/propostas", status_code=201)
def criar_proposta(dados: dict = Body(...), db: Session = Depends(get_db), _=Depends(get_current_user)):
    p = PropProposta(tipo=dados.get("tipo") or "venda", status=dados.get("status") or "rascunho",
                     template_id=dados.get("template_id"), empresa_id=dados.get("empresa_id"),
                     pessoa_id=dados.get("pessoa_id"), card_id=dados.get("card_id"), data=dados.get("data"),
                     validade_dias=dados.get("validade_dias"), texto_comercial=dados.get("texto_comercial") or {},
                     moeda_padrao=dados.get("moeda_padrao") or "BRL", blocos=dados.get("blocos") or {})
    db.add(p); db.flush()
    _apply_grupos(db, p.id, dados.get("grupos"))
    db.commit(); db.refresh(p)
    return _proposta_out(p)


@router.patch("/propostas/{pid}")
def atualizar_proposta(pid: UUID, dados: dict = Body(...), db: Session = Depends(get_db), _=Depends(get_current_user)):
    p = db.query(PropProposta).filter(PropProposta.id == pid).first()
    if not p: raise HTTPException(404, "Proposta nao encontrada")
    for k in ("tipo", "status", "template_id", "empresa_id", "pessoa_id", "card_id", "data", "validade_dias",
              "texto_comercial", "moeda_padrao", "blocos"):
        if k in dados: setattr(p, k, dados[k])
    if "grupos" in dados: _apply_grupos(db, pid, dados["grupos"])
    db.commit(); db.refresh(p)
    return _proposta_out(p)


@router.delete("/propostas/{pid}", status_code=204)
def deletar_proposta(pid: UUID, db: Session = Depends(get_db), _=Depends(get_current_user)):
    p = db.query(PropProposta).filter(PropProposta.id == pid).first()
    if p: db.delete(p); db.commit()


# ---------- Funil / Oportunidades ----------
def _username(u):
    if isinstance(u, dict): return u.get("username") or u.get("email") or "sistema"
    return str(u) if u else "sistema"


def _int_out(i):
    return {"id": str(i.id), "tipo": i.tipo, "texto": i.texto, "usuario": i.usuario,
            "anexos": i.anexos or [], "data_hora": i.data_hora.isoformat() if i.data_hora else None}


def _funil_out(f):
    return {"id": str(f.id), "nome": f.nome, "etapas": f.etapas or [], "ativo": f.ativo, "ordem": f.ordem or 0}


def _op_out(o):
    dias = (datetime.utcnow() - o.data_entrada_etapa).days if o.data_entrada_etapa else None
    return {"id": str(o.id), "funil_id": str(o.funil_id), "numero": o.numero, "etapa": o.etapa, "titulo": o.titulo,
            "empresa_id": str(o.empresa_id) if o.empresa_id else None,
            "pessoa_id": str(o.pessoa_id) if o.pessoa_id else None,
            "vendedor": o.vendedor, "marcadores": o.marcadores or [], "origem": o.origem, "tipo": o.tipo,
            "farol": o.farol or "frio", "sinaleiro": o.sinaleiro or "red", "ordem": o.ordem or 0,
            "arquivado": o.arquivado, "data_tarefa": o.data_tarefa.isoformat() if o.data_tarefa else None,
            "dias_etapa": dias, "criado_em": o.criado_em.isoformat() if o.criado_em else None}


@router.get("/funis")
def listar_funis(db: Session = Depends(get_db), _=Depends(get_current_user)):
    return [_funil_out(f) for f in db.query(Funil).filter(Funil.ativo == True).order_by(Funil.ordem, Funil.nome).all()]


@router.post("/funis", status_code=201)
def criar_funil(dados: dict = Body(...), db: Session = Depends(get_db), _=Depends(get_current_user)):
    f = Funil(nome=dados.get("nome") or "Novo funil", etapas=dados.get("etapas") or [], ordem=int(dados.get("ordem") or 0))
    db.add(f); db.commit(); db.refresh(f)
    return _funil_out(f)


@router.patch("/funis/{fid}")
def atualizar_funil(fid: UUID, dados: dict = Body(...), db: Session = Depends(get_db), _=Depends(get_current_user)):
    f = db.query(Funil).filter(Funil.id == fid).first()
    if not f: raise HTTPException(404, "Funil nao encontrado")
    for k in ("nome", "etapas", "ativo", "ordem"):
        if k in dados: setattr(f, k, dados[k])
    db.commit(); db.refresh(f)
    return _funil_out(f)


@router.delete("/funis/{fid}", status_code=204)
def deletar_funil(fid: UUID, db: Session = Depends(get_db), _=Depends(get_current_user)):
    f = db.query(Funil).filter(Funil.id == fid).first()
    if f: db.delete(f); db.commit()


@router.post("/funis/{fid}/duplicar", status_code=201)
def duplicar_funil(fid: UUID, dados: dict = Body(default={}), db: Session = Depends(get_db), _=Depends(get_current_user)):
    f = db.query(Funil).filter(Funil.id == fid).first()
    if not f: raise HTTPException(404, "Funil nao encontrado")
    novo = Funil(nome=(dados or {}).get("nome") or (f.nome + " (cópia)"), etapas=list(f.etapas or []), ordem=(f.ordem or 0) + 1)
    db.add(novo); db.commit(); db.refresh(novo)
    return _funil_out(novo)


@router.get("/oportunidades")
def listar_oportunidades(funil_id: str | None = None, incluir_arquivadas: bool = False,
                         db: Session = Depends(get_db), _=Depends(get_current_user)):
    q = db.query(Oportunidade)
    if funil_id: q = q.filter(Oportunidade.funil_id == funil_id)
    if not incluir_arquivadas: q = q.filter(Oportunidade.arquivado == False)
    return [_op_out(o) for o in q.order_by(Oportunidade.ordem, Oportunidade.criado_em).all()]


@router.post("/oportunidades", status_code=201)
def criar_oportunidade(dados: dict = Body(...), db: Session = Depends(get_db), _=Depends(get_current_user)):
    maxn = db.query(func.max(Oportunidade.numero)).scalar() or 0
    dt = dados.get("data_tarefa")
    o = Oportunidade(funil_id=dados.get("funil_id"), etapa=dados.get("etapa"), titulo=dados.get("titulo"),
                     empresa_id=dados.get("empresa_id"), pessoa_id=dados.get("pessoa_id"),
                     vendedor=dados.get("vendedor"), marcadores=dados.get("marcadores") or [],
                     origem=dados.get("origem"), tipo=dados.get("tipo"), farol=dados.get("farol") or "frio",
                     numero=maxn + 1, data_entrada_etapa=datetime.utcnow(),
                     data_tarefa=date.fromisoformat(dt) if dt else None)
    db.add(o); db.commit(); db.refresh(o)
    return _op_out(o)


@router.patch("/oportunidades/{oid}")
def atualizar_oportunidade(oid: UUID, dados: dict = Body(...), db: Session = Depends(get_db), user=Depends(get_current_user)):
    o = db.query(Oportunidade).filter(Oportunidade.id == oid).first()
    if not o: raise HTTPException(404, "Oportunidade nao encontrada")
    if "etapa" in dados and dados["etapa"] and dados["etapa"] != o.etapa:
        o.data_entrada_etapa = datetime.utcnow()
        db.add(OpInteracao(oportunidade_id=o.id, tipo="sistema", texto='Movido de "' + str(o.etapa or "") + '" para "' + str(dados["etapa"]) + '"', usuario=_username(user)))
    if "data_tarefa" in dados:
        o.data_tarefa = date.fromisoformat(dados["data_tarefa"]) if dados.get("data_tarefa") else None
    for k in ("etapa", "titulo", "empresa_id", "pessoa_id", "vendedor", "marcadores", "origem", "tipo", "farol", "sinaleiro", "ordem", "arquivado"):
        if k in dados: setattr(o, k, dados[k])
    db.commit(); db.refresh(o)
    return _op_out(o)


@router.delete("/oportunidades/{oid}", status_code=204)
def deletar_oportunidade(oid: UUID, db: Session = Depends(get_db), _=Depends(get_current_user)):
    o = db.query(Oportunidade).filter(Oportunidade.id == oid).first()
    if o: db.delete(o); db.commit()


@router.get("/oportunidades/{oid}")
def obter_oportunidade(oid: UUID, db: Session = Depends(get_db), _=Depends(get_current_user)):
    o = db.query(Oportunidade).filter(Oportunidade.id == oid).first()
    if not o: raise HTTPException(404, "Oportunidade nao encontrada")
    ints = db.query(OpInteracao).filter(OpInteracao.oportunidade_id == oid).order_by(OpInteracao.data_hora.desc()).all()
    d = _op_out(o); d["interacoes"] = [_int_out(x) for x in ints]
    return d


@router.post("/oportunidades/{oid}/interacoes", status_code=201)
def add_interacao(oid: UUID, tipo: str = Form("nota"), texto: str = Form(""), arquivo: UploadFile = File(None),
                  db: Session = Depends(get_db), user=Depends(get_current_user)):
    o = db.query(Oportunidade).filter(Oportunidade.id == oid).first()
    if not o: raise HTTPException(404, "Oportunidade nao encontrada")
    anexos = []
    if arquivo is not None:
        try:
            anexos = [salvar_imagem(arquivo, settings.upload_dir, "opint")]
        except Exception:
            anexos = []
    i = OpInteracao(oportunidade_id=oid, tipo=tipo or "nota", texto=texto, usuario=_username(user), anexos=anexos)
    db.add(i); db.commit(); db.refresh(i)
    return _int_out(i)
