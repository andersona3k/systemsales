import calendar
import io
from datetime import date, timedelta
from typing import List
from uuid import UUID
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from fastapi import APIRouter, Depends, HTTPException, Body, UploadFile, File
from fastapi.responses import Response
from sqlalchemy import extract, func
from sqlalchemy.orm import Session
from ..database import get_db
from ..models.financeiro import (
    Venda, VendaItem, Produto, FinConfig, FinLancamento, FinParcela,
    FinCcLancamento, FinCcParcela,
    Forecast, ForecastItem,
)
from ..schemas.financeiro import (
    VendaIn, VendaOut, ProdutoIn, ProdutoOut,
    FinLancamentoIn, FinLancamentoOut, FinParcelaOut,
    FinCcLancamentoIn, FinCcLancamentoOut, FinCcParcelaOut, FinCcResumoOut,
)
from ..services.auth_service import get_current_user
from ..config import get_settings
from ..services.img_service import salvar_imagem

router = APIRouter(prefix="/api/fin", tags=["Financeiro"])
settings = get_settings()


@router.get("/vendas", response_model=List[VendaOut])
def listar_vendas(estagio: str | None = None, vendedor: str | None = None, cliente: str | None = None,
                  db: Session = Depends(get_db), _=Depends(get_current_user)):
    q = db.query(Venda)
    if estagio: q = q.filter(Venda.estagio == estagio)
    if vendedor: q = q.filter(Venda.vendedor == vendedor)
    if cliente: q = q.filter(Venda.cliente.ilike(f"%{cliente}%"))
    return q.order_by(Venda.criado_em.desc()).all()


@router.post("/vendas", response_model=VendaOut, status_code=201)
def criar_venda(dados: VendaIn, db: Session = Depends(get_db), _=Depends(get_current_user)):
    d = dados.model_dump(); itens = d.pop("itens", [])
    v = Venda(**d); db.add(v); db.flush()
    for it in itens:
        db.add(VendaItem(venda_id=v.id, **it))
    db.commit(); db.refresh(v)
    return v


@router.get("/vendas/{vid}", response_model=VendaOut)
def obter_venda(vid: UUID, db: Session = Depends(get_db), _=Depends(get_current_user)):
    v = db.query(Venda).filter(Venda.id == vid).first()
    if not v: raise HTTPException(404, "Venda não encontrada")
    return v


@router.patch("/vendas/{vid}", response_model=VendaOut)
def atualizar_venda(vid: UUID, dados: dict = Body(...), db: Session = Depends(get_db), _=Depends(get_current_user)):
    v = db.query(Venda).filter(Venda.id == vid).first()
    if not v: raise HTTPException(404, "Venda não encontrada")
    for k in ("id_lead","estagio","data_venda","vendedor","cliente","quem_fatura","mensal","anexos","comissao_pago_em","comissao_status","comissao_meses"):
        if k in dados: setattr(v, k, dados[k])
    if "itens" in dados:
        db.query(VendaItem).filter(VendaItem.venda_id == vid).delete()
        for it in dados["itens"]:
            db.add(VendaItem(venda_id=vid, grupo=it.get("grupo"), produto=it.get("produto"), detalhes=it.get("detalhes"),
                             quem_fatura=it.get("quem_fatura"), moeda=it.get("moeda") or "BRL",
                             valor=it.get("valor") or 0, parcelas=it.get("parcelas") or 1, dias_pagamento=it.get("dias_pagamento") or 0,
                             nf_numero=it.get("nf_numero"), nf_data=it.get("nf_data"), nf_valor=it.get("nf_valor"),
                             contrato=it.get("contrato"), campos_extras=it.get("campos_extras") or {}))
    db.commit(); db.refresh(v)
    return v


@router.delete("/vendas/{vid}", status_code=204)
def deletar_venda(vid: UUID, db: Session = Depends(get_db), _=Depends(get_current_user)):
    v = db.query(Venda).filter(Venda.id == vid).first()
    if not v: raise HTTPException(404, "Venda não encontrada")
    db.delete(v); db.commit()


@router.post("/vendas/{vid}/anexo", response_model=VendaOut)
def anexo_venda(vid: UUID, arquivo: UploadFile = File(...), db: Session = Depends(get_db), _=Depends(get_current_user)):
    v = db.query(Venda).filter(Venda.id == vid).first()
    if not v: raise HTTPException(404, "Venda não encontrada")
    rel = salvar_imagem(arquivo, settings.upload_dir, "venda")
    v.anexos = list(v.anexos or []) + [rel]
    db.commit(); db.refresh(v)
    return v


@router.delete("/vendas/{vid}/anexo", response_model=VendaOut)
def del_anexo(vid: UUID, dados: dict = Body(...), db: Session = Depends(get_db), _=Depends(get_current_user)):
    v = db.query(Venda).filter(Venda.id == vid).first()
    if not v: raise HTTPException(404, "Venda não encontrada")
    v.anexos = [u for u in (v.anexos or []) if u != dados.get("url")]
    db.commit(); db.refresh(v)
    return v


@router.get("/produtos", response_model=List[ProdutoOut])
def listar_produtos(db: Session = Depends(get_db), _=Depends(get_current_user)):
    return db.query(Produto).filter(Produto.ativo == True).order_by(Produto.grupo, Produto.nome).all()


@router.post("/produtos", response_model=ProdutoOut, status_code=201)
def criar_produto(dados: ProdutoIn, db: Session = Depends(get_db), _=Depends(get_current_user)):
    p = Produto(**dados.model_dump()); db.add(p); db.commit(); db.refresh(p); return p


@router.patch("/produtos/{pid}", response_model=ProdutoOut)
def atualizar_produto(pid: UUID, dados: dict = Body(...), db: Session = Depends(get_db), _=Depends(get_current_user)):
    p = db.query(Produto).filter(Produto.id == pid).first()
    if not p: raise HTTPException(404, "Produto não encontrado")
    for k in ("nome","grupo","ativo"):
        if k in dados: setattr(p, k, dados[k])
    db.commit(); db.refresh(p); return p


@router.delete("/produtos/{pid}", status_code=204)
def deletar_produto(pid: UUID, db: Session = Depends(get_db), _=Depends(get_current_user)):
    p = db.query(Produto).filter(Produto.id == pid).first()
    if not p: raise HTTPException(404, "Produto não encontrado")
    p.ativo = False; db.commit()


@router.get("/config/{chave}")
def obter_config(chave: str, db: Session = Depends(get_db), _=Depends(get_current_user)):
    c = db.query(FinConfig).filter(FinConfig.chave == chave).first()
    return {"chave": chave, "valor": c.valor if c else None}


@router.put("/config/{chave}")
def salvar_config(chave: str, dados: dict = Body(...), db: Session = Depends(get_db), _=Depends(get_current_user)):
    c = db.query(FinConfig).filter(FinConfig.chave == chave).first()
    if not c: c = FinConfig(chave=chave, valor=dados.get("valor")); db.add(c)
    else: c.valor = dados.get("valor")
    db.commit()
    return {"chave": chave, "valor": c.valor}


# ===== Finanças Empresa =====

def _add_months(d: date, n: int) -> date:
    mes = d.month - 1 + n
    ano = d.year + mes // 12
    mes = mes % 12 + 1
    dia = min(d.day, calendar.monthrange(ano, mes)[1])
    return date(ano, mes, dia)


RESPONSAVEIS = ("Anderson", "Sil", "Sophia", "Casa", "A3K")


def _status_parcela(p: FinParcela) -> str:
    if p.pago_em:
        return "pago"
    hoje = date.today()
    if p.vencimento < hoje:
        return "atrasado"
    if p.vencimento.year == hoje.year and p.vencimento.month == hoje.month:
        return "vencendo"
    return "planejado"


def _parcela_out(p: FinParcela) -> dict:
    l = p.lancamento
    juros = p.juros or 0
    desconto = p.desconto or 0
    return {
        "id": p.id, "lancamento_id": p.lancamento_id, "numero": p.numero,
        "vencimento": p.vencimento, "valor": p.valor, "juros": juros, "desconto": desconto,
        "valor_final": (p.valor or 0) + juros - desconto,
        "modo_pagamento": p.modo_pagamento,
        "observacao": p.observacao, "anexos": p.anexos or [],
        "pago_em": p.pago_em, "status": _status_parcela(p),
        "grupo": l.grupo if l else None, "categoria": l.categoria if l else None, "metodo": l.metodo if l else None,
        "conta": l.conta if l else None, "responsavel": l.responsavel if l else None,
        "descricao": l.descricao if l else None,
        "credor_pagador": l.credor_pagador if l else None,
        "total_parcelas": len(l.parcelas) if l else None,
    }


@router.post("/exportar-xlsx")
def exportar_xlsx_generico(dados: dict = Body(...), _=Depends(get_current_user)):
    """Gera um xlsx a partir de linhas/colunas já filtradas no cliente (reutilizável por qualquer tela)."""
    linhas = dados.get("linhas") or []
    colunas = dados.get("colunas") or []
    arquivo = dados.get("arquivo") or "exportacao.xlsx"
    if not arquivo.endswith(".xlsx"):
        arquivo += ".xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for col, c in enumerate(colunas, 1):
        cel = ws.cell(row=1, column=col, value=c.get("label") or c.get("chave") or "")
        cel.fill = fill
        cel.font = font
        cel.alignment = Alignment(horizontal="center")
    for row, linha in enumerate(linhas, 2):
        for col, c in enumerate(colunas, 1):
            ws.cell(row=row, column=col, value=linha.get(c.get("chave")))
    for col in ws.columns:
        largura = max((len(str(cel.value or "")) for cel in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(largura + 4, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    return Response(
        content=buffer.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={arquivo}"},
    )


@router.get("/financas-empresa", response_model=List[FinParcelaOut])
def listar_financas_empresa(grupo: str | None = None, categoria: str | None = None, metodo: str | None = None, status: str | None = None,
                             mes: int | None = None, ano: int | None = None,
                             db: Session = Depends(get_db), _=Depends(get_current_user)):
    q = db.query(FinParcela).join(FinLancamento)
    if grupo: q = q.filter(FinLancamento.grupo == grupo)
    if categoria: q = q.filter(FinLancamento.categoria == categoria)
    if metodo: q = q.filter(FinLancamento.metodo == metodo)
    if mes: q = q.filter(extract('month', FinParcela.vencimento) == mes)
    if ano: q = q.filter(extract('year', FinParcela.vencimento) == ano)
    out = [_parcela_out(p) for p in q.order_by(FinParcela.vencimento).all()]
    if status:
        out = [o for o in out if o["status"] == status]
    return out


@router.post("/financas-empresa", response_model=FinLancamentoOut, status_code=201)
def criar_financa_empresa(dados: FinLancamentoIn, db: Session = Depends(get_db), _=Depends(get_current_user)):
    if dados.grupo not in ("pessoal", "empresa"):
        raise HTTPException(400, "Grupo inválido")
    if dados.categoria not in ("despesa", "divida", "receita", "investimento", "cartao", "dinheiro"):
        raise HTTPException(400, "Categoria inválida")
    if dados.responsavel and dados.responsavel not in RESPONSAVEIS:
        raise HTTPException(400, "Responsável inválido")
    if dados.metodo not in ("mensal", "financiamento", "pontual"):
        raise HTTPException(400, "Método inválido")
    if dados.metodo == "financiamento" and not dados.numero_parcelas:
        raise HTTPException(400, "Financiamento exige número de parcelas")

    d = dados.model_dump()
    modo_pagamento = d.pop("modo_pagamento", None)
    l = FinLancamento(**d)
    db.add(l); db.flush()

    qtd = 1 if dados.metodo == "pontual" else (12 if dados.metodo == "mensal" else dados.numero_parcelas)
    for i in range(qtd):
        venc = dados.data_inicio if i == 0 else _add_months(dados.data_inicio, i)
        db.add(FinParcela(lancamento_id=l.id, numero=i + 1, vencimento=venc, valor=dados.valor, modo_pagamento=modo_pagamento))
    db.commit(); db.refresh(l)
    return l


@router.patch("/financas-empresa/lancamentos/{lid}", response_model=FinLancamentoOut)
def atualizar_lancamento_empresa(lid: UUID, dados: dict = Body(...), db: Session = Depends(get_db), _=Depends(get_current_user)):
    l = db.query(FinLancamento).filter(FinLancamento.id == lid).first()
    if not l: raise HTTPException(404, "Lançamento não encontrado")
    if "grupo" in dados and dados["grupo"] not in ("pessoal", "empresa"):
        raise HTTPException(400, "Grupo inválido")
    if "categoria" in dados and dados["categoria"] not in ("despesa", "divida", "receita", "investimento", "cartao", "dinheiro"):
        raise HTTPException(400, "Categoria inválida")
    if dados.get("responsavel") and dados["responsavel"] not in RESPONSAVEIS:
        raise HTTPException(400, "Responsável inválido")
    for k in ("grupo", "categoria", "conta", "responsavel", "descricao", "credor_pagador", "valor"):
        if k in dados: setattr(l, k, dados[k])
    db.commit(); db.refresh(l)
    return l


@router.delete("/financas-empresa/lancamentos/{lid}", status_code=204)
def deletar_lancamento_empresa(lid: UUID, db: Session = Depends(get_db), _=Depends(get_current_user)):
    l = db.query(FinLancamento).filter(FinLancamento.id == lid).first()
    if not l: raise HTTPException(404, "Lançamento não encontrado")
    db.delete(l); db.commit()


@router.patch("/financas-empresa/parcelas/{pid}", response_model=FinParcelaOut)
def atualizar_parcela_empresa(pid: UUID, dados: dict = Body(...), db: Session = Depends(get_db), _=Depends(get_current_user)):
    p = db.query(FinParcela).filter(FinParcela.id == pid).first()
    if not p: raise HTTPException(404, "Parcela não encontrada")
    if "vencimento" in dados: p.vencimento = dados["vencimento"]
    if "valor" in dados: p.valor = dados["valor"]
    if "pago_em" in dados: p.pago_em = dados["pago_em"]
    if "juros" in dados: p.juros = max(0, dados["juros"] or 0)
    if "desconto" in dados: p.desconto = max(0, dados["desconto"] or 0)
    if "modo_pagamento" in dados: p.modo_pagamento = dados["modo_pagamento"]
    if "observacao" in dados: p.observacao = dados["observacao"]
    db.commit(); db.refresh(p)
    return _parcela_out(p)


@router.delete("/financas-empresa/parcelas/{pid}", status_code=204)
def deletar_parcela_empresa(pid: UUID, db: Session = Depends(get_db), _=Depends(get_current_user)):
    p = db.query(FinParcela).filter(FinParcela.id == pid).first()
    if not p: raise HTTPException(404, "Parcela não encontrada")
    lid = p.lancamento_id
    db.delete(p); db.flush()
    restantes = db.query(FinParcela).filter(FinParcela.lancamento_id == lid).count()
    if restantes == 0:
        l = db.query(FinLancamento).filter(FinLancamento.id == lid).first()
        if l: db.delete(l)
    db.commit()


@router.post("/financas-empresa/parcelas/{pid}/anexo", response_model=FinParcelaOut)
def anexo_parcela_empresa(pid: UUID, arquivo: UploadFile = File(...), db: Session = Depends(get_db), _=Depends(get_current_user)):
    p = db.query(FinParcela).filter(FinParcela.id == pid).first()
    if not p: raise HTTPException(404, "Parcela não encontrada")
    rel = salvar_imagem(arquivo, settings.upload_dir, "comprovante")
    p.anexos = list(p.anexos or []) + [rel]
    db.commit(); db.refresh(p)
    return _parcela_out(p)


@router.delete("/financas-empresa/parcelas/{pid}/anexo", response_model=FinParcelaOut)
def del_anexo_parcela_empresa(pid: UUID, dados: dict = Body(...), db: Session = Depends(get_db), _=Depends(get_current_user)):
    p = db.query(FinParcela).filter(FinParcela.id == pid).first()
    if not p: raise HTTPException(404, "Parcela não encontrada")
    p.anexos = [u for u in (p.anexos or []) if u != dados.get("url")]
    db.commit(); db.refresh(p)
    return _parcela_out(p)


# ===== Cartão de Crédito =====

CARTOES = {
    "pix": {"nome": "Pix", "dia": None},
    "dinheiro": {"nome": "Dinheiro", "dia": None},
    "debito": {"nome": "Débito", "dia": None},
    "nubank_gesser": {"nome": "Nubank Gesser", "dia": 22},
    "santander": {"nome": "Santander", "dia": 16},
    "mercado_pago": {"nome": "Mercado Pago", "dia": 17},
    "nubank_sil": {"nome": "Nubank Sil", "dia": 17},
    "caixa": {"nome": "Caixa", "dia": 20},
}


def _venc_fatura_base(cartao_dia: int | None, data_compra: date) -> date:
    if cartao_dia is None:
        return date(data_compra.year, data_compra.month, 1)
    venc_mes = date(data_compra.year, data_compra.month, cartao_dia)
    fechamento = venc_mes - timedelta(days=10)
    if data_compra <= fechamento:
        return venc_mes
    return _add_months(venc_mes, 1)


def _meses_ate_dezembro(data_compra: date) -> int:
    return 12 - data_compra.month + 1


def _venc_fatura(cartao_dia: int, data_compra: date, numero: int) -> date:
    base = _venc_fatura_base(cartao_dia, data_compra)
    return base if numero <= 1 else _add_months(base, numero - 1)


def _cc_parcela_out(p: FinCcParcela) -> dict:
    l = p.lancamento
    info = CARTOES.get(l.cartao) if l else None
    dia = info["dia"] if info else 1
    venc_fatura = _venc_fatura(dia, l.data_compra, p.numero or 1) if l else None
    return {
        "id": p.id, "lancamento_id": p.lancamento_id, "numero": p.numero,
        "valor": p.valor, "observacao": p.observacao, "anexos": p.anexos or [],
        "vencimento_fatura": venc_fatura,
        "grupo": l.grupo if l else None, "categoria": l.categoria if l else None,
        "cartao": l.cartao if l else None, "metodo": l.metodo if l else None,
        "metodo_pg": l.metodo_pg if l else None,
        "conta": l.conta if l else None, "sub_conta": l.sub_conta if l else None,
        "responsavel": l.responsavel if l else None,
        "descricao": l.descricao if l else None,
        "credor_pagador": l.credor_pagador if l else None,
        "data_compra": l.data_compra if l else None,
        "total_parcelas": len(l.parcelas) if l else None,
    }


@router.get("/financas-cartao", response_model=List[FinCcParcelaOut])
def listar_financas_cartao(grupo: str | None = None, categoria: str | None = None, cartao: str | None = None,
                            db: Session = Depends(get_db), _=Depends(get_current_user)):
    q = db.query(FinCcParcela).join(FinCcLancamento)
    if grupo: q = q.filter(FinCcLancamento.grupo == grupo)
    if categoria: q = q.filter(FinCcLancamento.categoria == categoria)
    if cartao: q = q.filter(FinCcLancamento.cartao == cartao)
    parcelas = q.order_by(FinCcLancamento.data_compra).all()
    return [_cc_parcela_out(p) for p in parcelas]


METODO_PG_PARA_MODO_PAGAMENTO = {"cartao": "credito", "pix": "pix", "dinheiro": "dinheiro", "debito": "debito_automatico"}


def _sync_compra_avista_controle(dados: FinCcLancamentoIn, cartao_final: str, db: Session):
    """Compra à vista (despesa/consumo) reflete automaticamente em 1 lançamento "Compras" no Controle financeiro,
    acumulando no lançamento já existente do mesmo mês/ano com o mesmo Grupo+Responsável+Categoria+Conta."""
    if dados.metodo != "avista" or dados.categoria not in ("despesa", "consumo"):
        return
    modo_pagamento = METODO_PG_PARA_MODO_PAGAMENTO.get(dados.metodo_pg)
    if dados.metodo_pg == "cartao":
        info = CARTOES.get(cartao_final)
        dia = info["dia"] if info else None
        data_ref = _venc_fatura_base(dia, dados.data_compra)
        pago_em = None
    else:
        data_ref = dados.data_compra
        pago_em = dados.data_compra

    match = (
        db.query(FinLancamento)
        .filter(
            FinLancamento.grupo == dados.grupo,
            FinLancamento.categoria == dados.categoria,
            FinLancamento.metodo == "compras",
            FinLancamento.responsavel == dados.responsavel,
            FinLancamento.conta == dados.conta,
            extract("month", FinLancamento.data_inicio) == data_ref.month,
            extract("year", FinLancamento.data_inicio) == data_ref.year,
        )
        .first()
    )
    if match:
        match.valor = (match.valor or 0) + dados.valor
        parcela = db.query(FinParcela).filter(FinParcela.lancamento_id == match.id).first()
        if parcela:
            parcela.valor = (parcela.valor or 0) + dados.valor
            parcela.modo_pagamento = modo_pagamento
    else:
        novo = FinLancamento(
            grupo=dados.grupo, categoria=dados.categoria, metodo="compras",
            conta=dados.conta, responsavel=dados.responsavel,
            descricao=dados.descricao, credor_pagador=dados.credor_pagador,
            valor=dados.valor, data_inicio=data_ref,
        )
        db.add(novo); db.flush()
        db.add(FinParcela(lancamento_id=novo.id, numero=1, vencimento=data_ref, valor=dados.valor, pago_em=pago_em, modo_pagamento=modo_pagamento))


@router.post("/financas-cartao", response_model=FinCcLancamentoOut, status_code=201)
def criar_financa_cartao(dados: FinCcLancamentoIn, db: Session = Depends(get_db), _=Depends(get_current_user)):
    if dados.grupo not in ("pessoal", "empresa"):
        raise HTTPException(400, "Grupo inválido")
    if dados.categoria not in ("despesa", "consumo", "assinatura"):
        raise HTTPException(400, "Categoria inválida")
    if dados.metodo_pg not in ("cartao", "dinheiro", "pix", "debito"):
        raise HTTPException(400, "Método PG inválido")
    if dados.responsavel and dados.responsavel not in RESPONSAVEIS:
        raise HTTPException(400, "Responsável inválido")
    cartao_final = dados.cartao if dados.metodo_pg == "cartao" else dados.metodo_pg
    if cartao_final not in CARTOES:
        raise HTTPException(400, "Cartão inválido")
    if dados.metodo not in ("avista", "parcelado", "recorrente"):
        raise HTTPException(400, "Forma de pagamento inválida")
    if dados.metodo == "parcelado" and not dados.numero_parcelas:
        raise HTTPException(400, "Parcelado exige número de parcelas")

    d = dados.model_dump()
    valores_parcelas = d.pop("valores_parcelas", None)
    d["cartao"] = cartao_final
    l = FinCcLancamento(**d)
    db.add(l); db.flush()

    if dados.metodo == "avista":
        qtd = 1
    elif dados.metodo == "parcelado":
        qtd = dados.numero_parcelas
    else:
        qtd = _meses_ate_dezembro(dados.data_compra)

    if dados.metodo == "parcelado" and valores_parcelas and len(valores_parcelas) == qtd:
        valores = valores_parcelas
    elif dados.metodo == "parcelado":
        base = round(dados.valor / qtd, 2)
        valores = [base] * (qtd - 1) + [round(dados.valor - base * (qtd - 1), 2)]
    else:
        valores = [dados.valor] * qtd
    for i in range(qtd):
        db.add(FinCcParcela(lancamento_id=l.id, numero=i + 1, valor=valores[i]))
    _sync_compra_avista_controle(dados, cartao_final, db)
    db.commit(); db.refresh(l)
    return l


@router.patch("/financas-cartao/lancamentos/{lid}", response_model=FinCcLancamentoOut)
def atualizar_lancamento_cartao(lid: UUID, dados: dict = Body(...), db: Session = Depends(get_db), _=Depends(get_current_user)):
    l = db.query(FinCcLancamento).filter(FinCcLancamento.id == lid).first()
    if not l: raise HTTPException(404, "Lançamento não encontrado")
    if "grupo" in dados and dados["grupo"] not in ("pessoal", "empresa"):
        raise HTTPException(400, "Grupo inválido")
    if "categoria" in dados and dados["categoria"] not in ("despesa", "consumo", "assinatura"):
        raise HTTPException(400, "Categoria inválida")
    if "metodo_pg" in dados and dados["metodo_pg"] not in ("cartao", "dinheiro", "pix", "debito"):
        raise HTTPException(400, "Método PG inválido")
    if "cartao" in dados and dados["cartao"] not in CARTOES:
        raise HTTPException(400, "Cartão inválido")
    if dados.get("responsavel") and dados["responsavel"] not in RESPONSAVEIS:
        raise HTTPException(400, "Responsável inválido")
    for k in ("grupo", "categoria", "cartao", "conta", "sub_conta", "responsavel", "metodo_pg", "descricao", "credor_pagador", "data_compra"):
        if k in dados: setattr(l, k, dados[k])
    db.commit(); db.refresh(l)
    return l


@router.delete("/financas-cartao/lancamentos/{lid}", status_code=204)
def deletar_lancamento_cartao(lid: UUID, db: Session = Depends(get_db), _=Depends(get_current_user)):
    l = db.query(FinCcLancamento).filter(FinCcLancamento.id == lid).first()
    if not l: raise HTTPException(404, "Lançamento não encontrado")
    db.delete(l); db.commit()


@router.patch("/financas-cartao/parcelas/{pid}", response_model=FinCcParcelaOut)
def atualizar_parcela_cartao(pid: UUID, dados: dict = Body(...), db: Session = Depends(get_db), _=Depends(get_current_user)):
    p = db.query(FinCcParcela).filter(FinCcParcela.id == pid).first()
    if not p: raise HTTPException(404, "Parcela não encontrada")
    if "valor" in dados: p.valor = dados["valor"]
    if "observacao" in dados: p.observacao = dados["observacao"]
    db.commit(); db.refresh(p)
    return _cc_parcela_out(p)


@router.delete("/financas-cartao/parcelas/{pid}", status_code=204)
def deletar_parcela_cartao(pid: UUID, db: Session = Depends(get_db), _=Depends(get_current_user)):
    p = db.query(FinCcParcela).filter(FinCcParcela.id == pid).first()
    if not p: raise HTTPException(404, "Parcela não encontrada")
    lid = p.lancamento_id
    db.delete(p); db.flush()
    restantes = db.query(FinCcParcela).filter(FinCcParcela.lancamento_id == lid).count()
    if restantes == 0:
        l = db.query(FinCcLancamento).filter(FinCcLancamento.id == lid).first()
        if l: db.delete(l)
    db.commit()


@router.post("/financas-cartao/parcelas/{pid}/anexo", response_model=FinCcParcelaOut)
def anexo_parcela_cartao(pid: UUID, arquivo: UploadFile = File(...), db: Session = Depends(get_db), _=Depends(get_current_user)):
    p = db.query(FinCcParcela).filter(FinCcParcela.id == pid).first()
    if not p: raise HTTPException(404, "Parcela não encontrada")
    rel = salvar_imagem(arquivo, settings.upload_dir, "cc_comprovante")
    p.anexos = list(p.anexos or []) + [rel]
    db.commit(); db.refresh(p)
    return _cc_parcela_out(p)


@router.delete("/financas-cartao/parcelas/{pid}/anexo", response_model=FinCcParcelaOut)
def del_anexo_parcela_cartao(pid: UUID, dados: dict = Body(...), db: Session = Depends(get_db), _=Depends(get_current_user)):
    p = db.query(FinCcParcela).filter(FinCcParcela.id == pid).first()
    if not p: raise HTTPException(404, "Parcela não encontrada")
    p.anexos = [u for u in (p.anexos or []) if u != dados.get("url")]
    db.commit(); db.refresh(p)
    return _cc_parcela_out(p)


@router.get("/financas-cartao/resumo", response_model=List[FinCcResumoOut])
def resumo_cartao(db: Session = Depends(get_db), _=Depends(get_current_user)):
    hoje = date.today()
    out = []
    for chave, info in CARTOES.items():
        fatura_aberta = _venc_fatura_base(info["dia"], hoje)
        parcelas = db.query(FinCcParcela).join(FinCcLancamento).filter(FinCcLancamento.cartao == chave).all()
        total = 0.0
        for p in parcelas:
            venc = _venc_fatura(info["dia"], p.lancamento.data_compra, p.numero or 1)
            if venc == fatura_aberta:
                total += p.valor or 0
        out.append({
            "cartao": chave, "nome": info["nome"], "vencimento_fatura": fatura_aberta, "total": total,
        })
    return out


# ===== FORECAST (Forcast) =====
def _fc_item_out(it):
    return {"id": str(it.id), "produto": it.produto, "classe": it.classe,
            "tipo_linha": it.tipo_linha or "Projeto", "quantidade": it.quantidade or 0,
            "valor_custo": it.valor_custo or 0, "valor_venda": it.valor_venda or 0,
            "meses": it.meses or 1, "ordem": it.ordem or 0}


def _fc_out(f):
    return {"id": str(f.id), "id_lead": f.id_lead, "cliente": f.cliente,
            "status": f.status or "Forcast", "previsao_fechamento": f.previsao_fechamento,
            "pct_fechamento": f.pct_fechamento or 0, "tipo": f.tipo,
            "criado_em": f.criado_em.isoformat() if f.criado_em else None,
            "itens": [_fc_item_out(x) for x in f.itens]}


def _fc_apply_itens(db, fid, itens):
    db.query(ForecastItem).filter(ForecastItem.forecast_id == fid).delete()
    for i, it in enumerate(itens or []):
        db.add(ForecastItem(
            forecast_id=fid, produto=it.get("produto"), classe=it.get("classe"),
            tipo_linha=it.get("tipo_linha") or "Projeto",
            quantidade=float(it.get("quantidade") or 0),
            valor_custo=float(it.get("valor_custo") or 0),
            valor_venda=float(it.get("valor_venda") or 0),
            meses=int(it.get("meses") or 1),
            ordem=it.get("ordem") if it.get("ordem") is not None else i))


@router.get("/forecast")
def listar_forecast(ano: str | None = None, db: Session = Depends(get_db), _=Depends(get_current_user)):
    q = db.query(Forecast)
    if ano:
        q = q.filter(Forecast.previsao_fechamento.like(f"{ano}-%"))
    return [_fc_out(f) for f in q.order_by(Forecast.criado_em.desc()).all()]


@router.post("/forecast", status_code=201)
def criar_forecast(dados: dict = Body(...), db: Session = Depends(get_db), _=Depends(get_current_user)):
    f = Forecast(id_lead=dados.get("id_lead"), cliente=dados.get("cliente"),
                 status=dados.get("status") or "Forcast",
                 previsao_fechamento=dados.get("previsao_fechamento"),
                 pct_fechamento=int(dados.get("pct_fechamento") or 0), tipo=dados.get("tipo"))
    db.add(f); db.flush()
    _fc_apply_itens(db, f.id, dados.get("itens"))
    db.commit(); db.refresh(f)
    return _fc_out(f)


@router.patch("/forecast/{fid}")
def atualizar_forecast(fid: UUID, dados: dict = Body(...), db: Session = Depends(get_db), _=Depends(get_current_user)):
    f = db.query(Forecast).filter(Forecast.id == fid).first()
    if not f: raise HTTPException(404, "Forecast nao encontrado")
    for k in ("id_lead", "cliente", "status", "previsao_fechamento", "tipo"):
        if k in dados: setattr(f, k, dados[k])
    if "pct_fechamento" in dados: f.pct_fechamento = int(dados.get("pct_fechamento") or 0)
    if "itens" in dados: _fc_apply_itens(db, fid, dados["itens"])
    db.commit(); db.refresh(f)
    return _fc_out(f)


@router.delete("/forecast/{fid}", status_code=204)
def deletar_forecast(fid: UUID, db: Session = Depends(get_db), _=Depends(get_current_user)):
    f = db.query(Forecast).filter(Forecast.id == fid).first()
    if not f: raise HTTPException(404, "Forecast nao encontrado")
    db.delete(f); db.commit()
