import csv
import io
import re
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

STATUS_LABEL = {
    "novo": "Novo",
    "lead_perdido": "Lead Perdido",
    "base_cliente": "Base Cliente",
    "ativar": "Ativar",
    "bloqueado": "Bloqueado",
    "arquivo_morto": "Arquivo Morto",
}

LINKEDIN_STATUS_LABEL = {"1": "1º grau", "nao_contato": "Não contato"}


def normalizar_status(valor):
    if not valor:
        return "novo"
    v = str(valor).strip().lower()
    for codigo, label in STATUS_LABEL.items():
        if v == codigo or v == label.lower():
            return codigo
    return "novo"


def normalizar_linkedin_status(valor):
    if not valor:
        return None
    v = str(valor).strip().lower()
    if v in ("1", "1º grau", "1o grau", "primeiro grau"):
        return "1"
    if v in ("nao_contato", "não contato", "nao contato"):
        return "nao_contato"
    return None


def telefone_valido(valor):
    if not valor:
        return False
    digitos = re.sub(r"\D", "", str(valor))
    return len(digitos) >= 10


def ler_planilha(conteudo: bytes, filename: str) -> list[dict]:
    nome = (filename or "").lower()
    if nome.endswith(".xlsx"):
        wb = openpyxl.load_workbook(io.BytesIO(conteudo), read_only=True, data_only=True)
        ws = wb.active
        linhas = list(ws.iter_rows(values_only=True))
        if not linhas:
            return []
        cabecalho = [str(c).strip() if c is not None else "" for c in linhas[0]]
        resultado = []
        for row in linhas[1:]:
            if all(c is None or str(c).strip() == "" for c in row):
                continue
            item = {}
            for i, chave in enumerate(cabecalho):
                if not chave:
                    continue
                v = row[i] if i < len(row) else None
                item[chave] = str(v).strip() if v is not None else ""
            resultado.append(item)
        return resultado
    texto = conteudo.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(texto))
    return [dict(row) for row in reader if any((v or "").strip() for v in row.values())]


def _gerar_csv(objetos, campos, labels, valor_fn) -> bytes:
    output = io.StringIO()
    writer = csv.writer(output, quoting=csv.QUOTE_ALL)
    writer.writerow([labels[c] for c in campos])
    for obj in objetos:
        writer.writerow([valor_fn(obj, c) for c in campos])
    return output.getvalue().encode("utf-8-sig")


def _gerar_xlsx(objetos, campos, labels, valor_fn) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    font = Font(color="FFFFFF", bold=True)
    for col, c in enumerate(campos, 1):
        cel = ws.cell(row=1, column=col, value=labels[c])
        cel.fill = fill
        cel.font = font
        cel.alignment = Alignment(horizontal="center")
    for row, obj in enumerate(objetos, 2):
        for col, c in enumerate(campos, 1):
            ws.cell(row=row, column=col, value=valor_fn(obj, c))
    for col in ws.columns:
        largura = max(len(str(cel.value or "")) for cel in col)
        ws.column_dimensions[col[0].column_letter].width = min(largura + 4, 50)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


EMPRESA_CAMPOS = ["nome", "cnpj", "segmento", "estado", "cidade", "website", "instagram", "linkedin_url", "origem", "status", "criado_em"]
EMPRESA_LABELS = {
    "nome": "Nome", "cnpj": "CNPJ", "segmento": "Segmento", "estado": "Estado", "cidade": "Cidade",
    "website": "Website", "instagram": "Instagram", "linkedin_url": "LinkedIn", "origem": "Origem",
    "status": "Status", "criado_em": "Data de Cadastro",
}


def _empresa_valor(e, campo):
    if campo == "status":
        return STATUS_LABEL.get(e.status or "novo", "Novo")
    if campo == "criado_em":
        return e.criado_em.strftime("%d/%m/%Y %H:%M") if e.criado_em else ""
    return getattr(e, campo, None) or ""


def exportar_empresas_csv(empresas):
    return _gerar_csv(empresas, EMPRESA_CAMPOS, EMPRESA_LABELS, _empresa_valor)


def exportar_empresas_xlsx(empresas):
    return _gerar_xlsx(empresas, EMPRESA_CAMPOS, EMPRESA_LABELS, _empresa_valor)


PESSOA_CAMPOS = ["nome", "empresa_nome", "cargo", "setor", "linkedin_status", "linkedin_url", "telefone", "email"]
PESSOA_LABELS = {
    "nome": "Nome", "empresa_nome": "Empresa", "cargo": "Cargo", "setor": "Setor",
    "linkedin_status": "LinkedIn Status", "linkedin_url": "LinkedIn URL", "telefone": "Telefone", "email": "Email",
}


def _pessoa_valor(p, campo):
    if campo == "empresa_nome":
        return p.empresa.nome if p.empresa else ""
    if campo == "linkedin_status":
        return LINKEDIN_STATUS_LABEL.get(p.linkedin_status or "", "")
    return getattr(p, campo, None) or ""


def exportar_pessoas_csv(pessoas):
    return _gerar_csv(pessoas, PESSOA_CAMPOS, PESSOA_LABELS, _pessoa_valor)


def exportar_pessoas_xlsx(pessoas):
    return _gerar_xlsx(pessoas, PESSOA_CAMPOS, PESSOA_LABELS, _pessoa_valor)
