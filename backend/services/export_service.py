import io
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
from typing import Any


CAMPOS_DISPONIVEIS = {
    "nome": "Nome",
    "cargo": "Cargo",
    "empresa_nome": "Empresa",
    "telefone1": "Telefone 1",
    "telefone2": "Telefone 2",
    "email": "E-mail",
    "website": "Website",
    "linkedin": "LinkedIn",
    "instagram": "Instagram",
    "endereco": "Endereço",
    "origem": "Origem",
    "notas": "Notas",
    "criado_em": "Data de Cadastro",
}

ORIGENS_LABEL = {
    "cartao_fisico": "Cartão Físico",
    "manual": "Manual",
    "feira_evento": "Feira/Evento",
    "indicacao": "Indicação",
    "outro": "Outro",
}


def _get_valor(contato: Any, campo: str, mapeamento: dict) -> str:
    valor = getattr(contato, campo, None)
    if campo == "empresa_nome":
        valor = contato.empresa.nome if contato.empresa else ""
    if campo == "origem" and valor:
        valor = ORIGENS_LABEL.get(str(valor), str(valor))
    if campo == "criado_em" and valor:
        valor = valor.strftime("%d/%m/%Y %H:%M")
    return str(valor) if valor is not None else ""


def exportar_csv(contatos: list, campos: list[str], mapeamento: dict[str, str]) -> bytes:
    output = io.StringIO()
    cabecalhos = [mapeamento.get(c, CAMPOS_DISPONIVEIS.get(c, c)) for c in campos]
    writer = csv.writer(output, quoting=csv.QUOTE_ALL)
    writer.writerow(cabecalhos)
    for contato in contatos:
        linha = [_get_valor(contato, c, mapeamento) for c in campos]
        writer.writerow(linha)
    return output.getvalue().encode("utf-8-sig")


def exportar_xlsx(contatos: list, campos: list[str], mapeamento: dict[str, str]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Contatos"

    cabecalho_fill = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
    cabecalho_font = Font(color="FFFFFF", bold=True)

    cabecalhos = [mapeamento.get(c, CAMPOS_DISPONIVEIS.get(c, c)) for c in campos]
    for col, cab in enumerate(cabecalhos, 1):
        cel = ws.cell(row=1, column=col, value=cab)
        cel.fill = cabecalho_fill
        cel.font = cabecalho_font
        cel.alignment = Alignment(horizontal="center")

    for row, contato in enumerate(contatos, 2):
        for col, campo in enumerate(campos, 1):
            ws.cell(row=row, column=col, value=_get_valor(contato, campo, mapeamento))

    for col in ws.columns:
        max_len = max(len(str(cel.value or "")) for cel in col)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def exportar_vcf(contatos: list) -> bytes:
    linhas = []
    for c in contatos:
        linhas.append("BEGIN:VCARD")
        linhas.append("VERSION:3.0")
        linhas.append(f"FN:{c.nome}")
        partes = c.nome.split(" ", 1)
        sobrenome = partes[1] if len(partes) > 1 else ""
        linhas.append(f"N:{sobrenome};{partes[0]};;;")
        empresa_nome = c.empresa.nome if c.empresa else ""
        if empresa_nome:
            linhas.append(f"ORG:{empresa_nome}")
        if c.cargo:
            linhas.append(f"TITLE:{c.cargo}")
        if c.telefone1:
            linhas.append(f"TEL;TYPE=WORK,VOICE:{c.telefone1}")
        if c.telefone2:
            linhas.append(f"TEL;TYPE=CELL,VOICE:{c.telefone2}")
        if c.email:
            linhas.append(f"EMAIL;TYPE=WORK:{c.email}")
        if c.website:
            linhas.append(f"URL:{c.website}")
        if c.linkedin:
            linhas.append(f"X-SOCIALPROFILE;TYPE=linkedin:{c.linkedin}")
        if c.endereco:
            linhas.append(f"ADR;TYPE=WORK:;;{c.endereco};;;;")
        if c.notas:
            linhas.append(f"NOTE:{c.notas}")
        linhas.append("END:VCARD")
        linhas.append("")
    return "\r\n".join(linhas).encode("utf-8")
